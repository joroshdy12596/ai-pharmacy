import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# قراءة ملف CSV
data = []
with open('price_anomalies_latest.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        data.append(row)

# إنشاء Workbook
wb = Workbook()
ws = wb.active
ws.title = "الأسعار الشاذة"

# إضافة رؤوس الأعمدة
headers = ['ID', 'اسم المنتج', 'سعر البيع الحالي', 'سعر الشراء', 'الكمية المباعة', 'الإيرادات', 'متوسط سعر البيع', 'متوسط الربح للوحدة', 'الأسباب']
ws.append(headers)

# تنسيق الرأس
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# إضافة البيانات
for row_data in data:
    avg_profit = float(row_data.get('avg_hist_profit_per_unit', 0) or 0)
    
    # تحديد اللون بناءً على الربح
    if avg_profit < 0:
        fill_color = "FFC7CE"  # أحمر فاتح
    elif avg_profit < 1:
        fill_color = "FFEB9C"  # أصفر فاتح
    else:
        fill_color = "C6EFCE"  # أخضر فاتح
    
    ws.append([
        row_data.get('id', ''),
        row_data.get('name', ''),
        float(row_data.get('current_price', 0) or 0),
        float(row_data.get('purchase_price', 0) or 0),
        int(row_data.get('qty_sold', 0) or 0),
        float(row_data.get('revenue', 0) or 0),
        float(row_data.get('avg_sold_price', 0) or 0) if row_data.get('avg_sold_price') else '',
        avg_profit,
        row_data.get('reasons', ''),
    ])
    
    # تلوين الصف بناءً على الربح
    current_row = ws.max_row
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for cell in ws[current_row]:
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right' if cell.column == 2 else 'center', vertical='center', wrap_text=True)

# تعديل عرض الأعمدة
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 30

# تجميد الرأس
ws.freeze_panes = 'A2'

# حفظ الملف
output_file = 'price_anomalies_2025.xlsx'
wb.save(output_file)
print(f'✓ تم إنشاء ملف Excel بنجاح: {output_file}')
print(f'✓ عدد المنتجات الشاذة: {len(data)}')
print()
print('ملخص المشاكل:')

# إحصائيات الأسباب
reason_count = {}
for row in data:
    reasons = row.get('reasons', '').split(';')
    for reason in reasons:
        reason = reason.strip()
        if reason:
            reason_count[reason] = reason_count.get(reason, 0) + 1

for reason, count in sorted(reason_count.items(), key=lambda x: -x[1]):
    print(f'  - {reason}: {count} منتج')

# إحصائيات الخسارة
negative_profit = [r for r in data if float(r.get('avg_hist_profit_per_unit', 0) or 0) < 0]
print()
print(f'⚠ المنتجات التي تبيع بخسارة: {len(negative_profit)} منتج')
